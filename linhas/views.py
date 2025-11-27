from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.utils import timezone
from datetime import timedelta
from .models import LoginAttempt
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.db import models

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def login_view(request):
    if request.method == 'POST' and request.POST.get('login_form'):
        username = request.POST.get('username')
        password = request.POST.get('password')
        ip_address = get_client_ip(request)
        
        print(f"Login attempt - Username: {username}, IP: {ip_address}")
        print(f"POST data: {request.POST}")
        
        # Check if user is blocked
        login_attempt = LoginAttempt.objects.filter(
            username=username,
            ip_address=ip_address,
            is_blocked=True
        ).first()

        if login_attempt and login_attempt.is_currently_blocked():
            remaining_time = login_attempt.blocked_until - timezone.now()
            minutes = int(remaining_time.total_seconds() / 60)
            return render(request, 'login.html', {
                'error': f'Conta bloqueada. Tente novamente em {minutes} minutos ou entre em contato com o suporte.',
                'show_support': True
            })

        user = authenticate(username=username, password=password)
        
        if user is not None:
            # Reset login attempts on successful login
            LoginAttempt.objects.filter(username=username, ip_address=ip_address).delete()
            login(request, user)
            return redirect('linhas:dashboard')
        else:
            # Track failed login attempt
            login_attempt = LoginAttempt.objects.filter(
                username=username,
                ip_address=ip_address,
                is_blocked=False
            ).first()

            print(f"Failed login - Current attempt: {login_attempt}")

            if login_attempt:
                login_attempt.attempt_count += 1
                print(f"Incrementing attempt count to: {login_attempt.attempt_count}")
                
                if login_attempt.attempt_count >= 3:
                    print("Blocking user - 3 or more attempts")
                    login_attempt.is_blocked = True
                    login_attempt.blocked_until = timezone.now() + timedelta(minutes=15)
                    login_attempt.save()
                    return render(request, 'login.html', {
                        'error': 'Conta bloqueada por 15 minutos devido a múltiplas tentativas incorretas.',
                        'show_support': True
                    })
                login_attempt.save()
                print(f"Saved attempt count: {login_attempt.attempt_count}")
                
                # Calcular tentativas restantes
                remaining_attempts = 3 - login_attempt.attempt_count
                return render(request, 'login.html', {
                    'error': f'Usuário ou senha incorretos. {remaining_attempts} tentativa(s) restante(s) antes do bloqueio.',
                    'remaining_attempts': remaining_attempts,
                    'attempt_count': login_attempt.attempt_count
                })
            else:
                print("Creating new login attempt record")
                login_attempt = LoginAttempt.objects.create(
                    username=username,
                    ip_address=ip_address
                )
                print(f"Created new attempt record: {login_attempt.id}")
                
                return render(request, 'login.html', {
                    'error': 'Usuário ou senha incorretos. 2 tentativa(s) restante(s) antes do bloqueio.',
                    'remaining_attempts': 2,
                    'attempt_count': 1
                })

    return render(request, 'login.html')

def listar_rps_cliente(request):
    cnpj = request.GET.get('cnpj', '')
    empresa = request.GET.get('empresa', '')
    rps = Linha.objects.filter(cnpj=cnpj, empresa=empresa).values_list('rp', flat=True).distinct()
    resultados = [rp for rp in rps if rp]
    return JsonResponse(resultados, safe=False)
def autocomplete_cnpj(request):
    termo = request.GET.get('q', '')
    linhas = Linha.objects.filter(cnpj__icontains=termo).values('cnpj', 'empresa').distinct()
    resultados = [
        {'cnpj': l['cnpj'], 'empresa': l['empresa']}
        for l in linhas[:10]
    ]
    return JsonResponse(resultados, safe=False)
from django.http import JsonResponse, HttpResponse
from .models import Linha, Protocolo
import csv

def buscar_cliente_por_cnpj(request):
    cnpj = request.GET.get('cnpj', '')
    linha = Linha.objects.filter(cnpj=cnpj).first()
    if linha:
        return JsonResponse({'empresa': linha.empresa})
    return JsonResponse({'empresa': ''})

def buscar_cnpj_api_externa(request):
    """
    View para buscar dados de CNPJ usando APIs externas (ReceitaWS e BrasilAPI)
    
    Usage: GET /linhas/buscar-cnpj-api/?cnpj=19131243000197
    """
    # Import seguro do requests para retornar erro amigável se ausente
    try:
        import requests
    except Exception as e:
        msg = 'Dependência ausente: requests'
        if settings.DEBUG:
            print('[CNPJ LOOKUP][ERRO IMPORT REQUESTS]', e)
        return JsonResponse({
            'success': False,
            'error': msg,
            'details': str(e) if settings.DEBUG else ''
        }, status=500)

    import re
    
    cnpj = request.GET.get('cnpj', '')
    
    if not cnpj:
        return JsonResponse({
            'success': False, 
            'error': 'CNPJ não informado'
        }, status=400)
    
    # Limpar CNPJ
    cnpj_limpo = re.sub(r'\D', '', cnpj)
    if settings.DEBUG:
        print(f'[CNPJ LOOKUP] Início busca - cnpj_recebido={cnpj} cnpj_limpo={cnpj_limpo}')
    
    if len(cnpj_limpo) != 14:
        return JsonResponse({
            'success': False,
            'error': 'CNPJ deve ter 14 dígitos'
        }, status=400)
    
    try:
        # Tentar BrasilAPI primeiro
        url_brasil = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_limpo}"
        if settings.DEBUG:
            print('[CNPJ LOOKUP] Consultando BrasilAPI:', url_brasil)
        response = requests.get(url_brasil, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            dados_padronizados = {
                'cnpj': data.get('cnpj', cnpj_limpo),
                'nome': data.get('razao_social', ''),
                'razao_social': data.get('razao_social', ''),
                'fantasia': data.get('nome_fantasia', ''),
                'situacao': data.get('descricao_situacao_cadastral', ''),
                'endereco': data.get('logradouro', ''),
                'numero': data.get('numero', ''),
                'complemento': data.get('complemento', ''),
                'bairro': data.get('bairro', ''),
                'municipio': data.get('municipio', ''),
                'uf': data.get('uf', ''),
                'cep': data.get('cep', ''),
                'telefone': data.get('ddd_telefone_1', ''),
                'email': data.get('email', ''),
                'data_abertura': data.get('data_inicio_atividade', ''),
                'fonte': 'BrasilAPI'
            }
            
            if settings.DEBUG:
                print('[CNPJ LOOKUP] BrasilAPI sucesso para', cnpj_limpo)
            return JsonResponse({
                'success': True,
                'dados': dados_padronizados
            })
        
    except Exception as e:
        if settings.DEBUG:
            print('[CNPJ LOOKUP][BrasilAPI][ERRO]', repr(e))
        # Fallback para ReceitaWS
    
    try:
        # Fallback ReceitaWS
        url_receita = f"https://www.receitaws.com.br/v1/cnpj/{cnpj_limpo}"
        if settings.DEBUG:
            print('[CNPJ LOOKUP] Consultando ReceitaWS:', url_receita)
        response = requests.get(url_receita, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            
            # Verificar se a API retornou erro
            if data.get("status") == "ERROR":
                return JsonResponse({
                    'success': False,
                    'error': data.get('message', 'Erro na consulta CNPJ')
                }, status=404)
            
            dados_padronizados = {
                'cnpj': data.get('cnpj', cnpj_limpo),
                'nome': data.get('nome', ''),
                'razao_social': data.get('nome', ''),
                'fantasia': data.get('fantasia', ''),
                'situacao': data.get('situacao', ''),
                'endereco': data.get('logradouro', ''),
                'numero': data.get('numero', ''),
                'complemento': data.get('complemento', ''),
                'bairro': data.get('bairro', ''),
                'municipio': data.get('municipio', ''),
                'uf': data.get('uf', ''),
                'cep': data.get('cep', ''),
                'telefone': data.get('telefone', ''),
                'email': data.get('email', ''),
                'data_abertura': data.get('abertura', ''),
                'fonte': 'ReceitaWS'
            }
            
            if settings.DEBUG:
                print('[CNPJ LOOKUP] ReceitaWS sucesso para', cnpj_limpo)
            return JsonResponse({
                'success': True,
                'dados': dados_padronizados
            })
        
    except Exception as e:
        if settings.DEBUG:
            print('[CNPJ LOOKUP][ReceitaWS][ERRO]', repr(e))
    
    if settings.DEBUG:
        print('[CNPJ LOOKUP] Não encontrado em nenhuma API para', cnpj_limpo)
    return JsonResponse({
        'success': False,
        'error': 'CNPJ não encontrado nas APIs externas'
    }, status=404)

def test_cnpj_api_view(request):
    """
    View para página de teste da API de CNPJ
    """
    return render(request, 'test_cnpj_api.html')

def debug_cnpj_button_view(request):
    """
    View para debug do botão CNPJ
    """
    return render(request, 'debug_cnpj_button.html')

def debug_cnpj_complete_view(request):
    """
    View para debug completo do CNPJ
    """
    return render(request, 'debug_cnpj_complete.html')

from django.http import JsonResponse
from .models import Linha

def buscar_linhas_estoque(request):
    termo = request.GET.get('q', '')
    empresa = request.GET.get('empresa', '')
    linhas = Linha.objects.filter(origem='ESTOQUE', empresa__icontains=empresa, numero__icontains=termo)
    resultados = [
        {'id': l.id, 'numero': l.numero, 'iccid': l.iccid}
        for l in linhas[:10]
    ]
    return JsonResponse(resultados, safe=False)


def buscar_empresas(request):
    """AJAX: retornar lista de empresas (empresa, cnpj) que contenham o termo q
    Retorna JSON: [{ 'empresa': 'Nome', 'cnpj': '00000000000000' }, ...]
    """
    q = request.GET.get('q', '').strip()
    # buscar por empresa parcialmente
    empresas_qs = Linha.objects.filter(empresa__icontains=q).values('empresa', 'cnpj').distinct()[:50]
    # deduplicate by empresa keeping first cnpj
    seen = set()
    results = []
    for row in empresas_qs:
        nome = row.get('empresa') or ''
        cnpj = row.get('cnpj') or ''
        if not nome:
            continue
        if nome in seen:
            continue
        seen.add(nome)
        results.append({'empresa': nome, 'cnpj': cnpj})
    return JsonResponse(results, safe=False)


def buscar_clientes(request):
    """AJAX: retornar lista de clientes existentes que contenham o termo q
    Retorna JSON: [{ 'id': 1, 'empresa': 'Nome', 'cnpj': '00000000000000', 'fantasia': '...', 'razao_social': '...'}, ...]
    """
    q = request.GET.get('q', '').strip()
    qs = Cliente.objects.all()
    if q:
        qs = qs.filter(models.Q(empresa__icontains=q) | models.Q(cnpj__icontains=q) | models.Q(fantasia__icontains=q) | models.Q(razao_social__icontains=q))
    results = []
    for c in qs[:50]:
        results.append({
            'id': c.id, 
            'empresa': c.empresa, 
            'cnpj': c.cnpj, 
            'fantasia': c.fantasia, 
            'razao_social': c.razao_social,
            'endereco_completo': c.endereco_completo or '',
            'contato': c.contato or '',
            'email': c.email or '',
            'telefone': c.telefone or '',
            'valor_taxa_manutencao': str(c.valor_taxa_manutencao) if c.valor_taxa_manutencao else '0.00'
        })
    return JsonResponse(results, safe=False)

@login_required
def buscar_cliente_completo(request):
    """AJAX: Buscar cliente por nome, CNPJ ou número de linha e retornar dados completos
    Permite busca por:
    - Nome da empresa/cliente
    - CNPJ
    - Número da linha (busca o cliente associado à linha)
    """
    from django.http import JsonResponse
    
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 2:
        return JsonResponse({'success': False, 'message': 'Digite pelo menos 2 caracteres'})
    
    # Remover caracteres especiais para busca de CNPJ/telefone
    q_digits = ''.join(filter(str.isdigit, q))
    
    cliente = None
    found_via = None
    
    # 1. Tentar buscar por número da linha primeiro
    if q_digits and len(q_digits) >= 8:  # Mínimo para ser um número de telefone
        try:
            linha = Linha.objects.filter(
                Q(numero__icontains=q) | Q(numero__icontains=q_digits)
            ).select_related('cliente').first()
            
            if linha and linha.cliente:
                cliente = linha.cliente
                found_via = f'linha {linha.numero}'
        except:
            pass
    
    # 2. Se não encontrou por linha, buscar diretamente no Cliente
    if not cliente:
        try:
            # Buscar por CNPJ (completo ou parcial)
            if q_digits and len(q_digits) >= 8:
                cliente = Cliente.objects.filter(
                    Q(cnpj__icontains=q_digits) | Q(cnpj__icontains=q)
                ).first()
                if cliente:
                    found_via = 'CNPJ'
            
            # Se não encontrou por CNPJ, buscar por nome/empresa
            if not cliente:
                cliente = Cliente.objects.filter(
                    Q(empresa__icontains=q) | 
                    Q(fantasia__icontains=q) |
                    Q(razao_social__icontains=q)
                ).first()
                if cliente:
                    found_via = 'nome da empresa'
        except:
            pass
    
    if not cliente:
        return JsonResponse({
            'success': False, 
            'message': f'Nenhum cliente encontrado para "{q}"'
        })
    
    # Retornar dados completos do cliente
    return JsonResponse({
        'success': True,
        'found_via': found_via,
        'cliente': {
            'id': cliente.id,
            'empresa': cliente.empresa or '',
            'cnpj': cliente.cnpj or '',
            'razao_social': cliente.razao_social or '',
            'fantasia': cliente.fantasia or '',
            'endereco_completo': cliente.endereco_completo or '',
            'contato': cliente.contato or '',
            'email': cliente.email or '',
            'telefone': cliente.telefone or '',
            'nome_dono': cliente.nome_dono or '',
            'cpf_dono': cliente.cpf_dono or '',
            'data_nascimento_dono': cliente.data_nascimento_dono.strftime('%d/%m/%Y') if cliente.data_nascimento_dono else '',
            'valor_taxa_manutencao': str(cliente.valor_taxa_manutencao) if cliente.valor_taxa_manutencao else '0.00'
        }
    })
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Linha, Protocolo, Cliente
from .forms import LinhaForm, BuscaLinhaForm, ClienteForm
from django.urls import reverse

@login_required
def lista_linhas(request):
    form_busca = BuscaLinhaForm(request.GET)
    linhas = Linha.objects.all()
    
    if form_busca.is_valid():
        busca = form_busca.cleaned_data.get('busca')
        operadora = form_busca.cleaned_data.get('operadora')
        tipo_plano = form_busca.cleaned_data.get('tipo_plano')
        ativa = form_busca.cleaned_data.get('ativa')
        
        if busca:
            linhas = linhas.filter(
                Q(numero__icontains=busca) |
                Q(empresa__icontains=busca) |
                Q(cnpj__icontains=busca) |
                Q(rp__icontains=busca)
            )
        
        if operadora:
            linhas = linhas.filter(operadora=operadora)
            
        if tipo_plano:
            linhas = linhas.filter(tipo_plano=tipo_plano)
            
        if ativa:
            if ativa == 'ativa':
                linhas = linhas.filter(ativa=True)
            elif ativa == 'inativa':
                linhas = linhas.filter(ativa=False)
            else:
                linhas = linhas.filter(acao__iexact=ativa)
    
    paginator = Paginator(linhas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form_busca': form_busca,
        'total_linhas': linhas.count(),
        'linhas_ativas': linhas.filter(ativa=True).count(),
        'linhas_inativas': linhas.filter(ativa=False).count(),
    }
    
    return render(request, 'linhas/lista_linhas.html', context)


@login_required
def cliente_novo(request):
    """Criar novo cliente (form simples)."""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            # Verifica se deve redirecionar para nova linha ou ficar na página
            next_action = request.POST.get('next_action', 'stay')
            
            if next_action == 'nova_linha':
                # Adiciona mensagem antes do redirect
                messages.success(request, f'Cliente "{cliente.empresa}" cadastrado com sucesso!')
                # Redireciona para a página de nova linha e já preenche empresa/cnpj via GET
                url = reverse('linhas:novalinha')
                params = f'?empresa={cliente.empresa}&cnpj={cliente.cnpj}'
                return redirect(url + params)
            else:
                # Fica na página atual, limpa o formulário para novo cadastro
                messages.success(request, f'Cliente "{cliente.empresa}" cadastrado com sucesso! Operação concluída.')
                form = ClienteForm()
                return render(request, 'linhas/novo_cliente.html', {
                    'form': form,
                    'cliente_salvo': cliente
                })
        else:
            # Criar mensagem de erro mais específica
            error_messages = []
            for field_name, field_errors in form.errors.items():
                field_label = form.fields[field_name].label or field_name
                for error in field_errors:
                    error_messages.append(f"{field_label}: {error}")
            
            if error_messages:
                messages.error(request, f'Erro ao cadastrar cliente: {"; ".join(error_messages[:3])}')
            else:
                messages.error(request, 'Erro ao cadastrar cliente. Verifique os dados informados.')
    else:
        form = ClienteForm()
    return render(request, 'linhas/novo_cliente.html', {'form': form})


@login_required
def cliente_create_ajax(request):
    """Receive AJAX POST to create a Cliente and return JSON response."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'success': False, 'errors': {'__all__': 'Método inválido.'}}, status=400)
    form = ClienteForm(request.POST)
    if form.is_valid():
        cliente = form.save()
        return JsonResponse({'success': True, 'cliente': {
            'id': cliente.id,
            'empresa': cliente.empresa,
            'cnpj': cliente.cnpj,
            'razao_social': cliente.razao_social,
            'fantasia': cliente.fantasia,
        }})
    else:
        # format errors as dict
        errors = {k: v for k, v in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)

@login_required
def nova_linha(request):
    if request.method == 'POST':
        form = LinhaForm(request.POST, request.FILES)
        observacoes_lateral = request.POST.get('observacoes_lateral', '')
        anexos = request.FILES.getlist('anexos')
        if form.is_valid():
            linha = form.save(commit=False)
            linha.criado_por = request.user
            # Adiciona observações da lateral se ação for TT ou PORTABILIDADE
            if linha.acao in ['TT', 'PORTABILIDADE'] and observacoes_lateral:
                if linha.observacoes:
                    linha.observacoes += '\n' + observacoes_lateral
                else:
                    linha.observacoes = observacoes_lateral
            linha.save()
            # Attempt to attach a Cliente if one exists with the same CNPJ or empresa
            try:
                import re
                cliente = None
                if linha.cnpj:
                    cnpj_digits = re.sub(r"\D", "", linha.cnpj)
                    cliente = Cliente.objects.filter(cnpj__icontains=cnpj_digits).first()
                if not cliente and linha.empresa:
                    cliente = Cliente.objects.filter(empresa__iexact=linha.empresa).first()
                if cliente:
                    linha.cliente = cliente
                    linha.save()
            except Exception:
                # don't block save on any lookup error
                pass
            # Salvar anexos (exemplo: salvar em pasta e registrar caminho, adaptar conforme modelo)
            for arquivo in anexos:
                with open(f'media/anexos/{arquivo.name}', 'wb+') as dest:
                    for chunk in arquivo.chunks():
                        dest.write(chunk)
            messages.success(request, f'Linha {linha.numero} cadastrada com sucesso!')
            return redirect('linhas:listalinhas')
    else:
        # Allow prefilling empresa/cnpj when redirected after creating a Cliente
        initial = {}
        empresa_prefill = request.GET.get('empresa')
        cnpj_prefill = request.GET.get('cnpj')
        if empresa_prefill:
            initial['empresa'] = empresa_prefill
        if cnpj_prefill:
            initial['cnpj'] = cnpj_prefill
        form = LinhaForm(initial=initial)
    return render(request, 'linhas/nova_linha.html', {'form': form})

@login_required
def editar_linha(request, pk):
    linha = get_object_or_404(Linha, pk=pk)
    
    if request.method == 'POST':
        form = LinhaForm(request.POST, instance=linha)
        if form.is_valid():
            form.save()
            messages.success(request, f'Linha {linha.numero} atualizada com sucesso!')
            return redirect('linhas:detalheslinha', pk=linha.pk)
    else:
        form = LinhaForm(instance=linha)
    
    return render(request, 'linhas/editar_linha.html', {'form': form, 'linha': linha})

@login_required
def detalhes_linha(request, pk):
    linha = get_object_or_404(Linha, pk=pk)
    return render(request, 'linhas/detalhes_linha.html', {'linha': linha})

@login_required
def excluir_linha(request, pk):
    linha = get_object_or_404(Linha, pk=pk)
    
    if request.method == 'POST':
        numero = linha.numero
        linha.delete()
        messages.success(request, f'Linha {numero} excluída com sucesso!')
        return redirect('linhas:listalinhas')
    
    return render(request, 'linhas/excluir_linha.html', {'linha': linha})


# Placeholder pages for top-level menu
@login_required
def dashboard(request):
    from django.db.models import Count, Avg
    from django.utils import timezone
    from datetime import timedelta

    # Compute billing cycle period: 19th -> 18th
    from datetime import datetime, date
    hoje = timezone.now().date()
    if hoje.day >= 19:
        ciclo_start = date(hoje.year, hoje.month, 19)
        # next month
        if hoje.month == 12:
            ciclo_end = date(hoje.year + 1, 1, 18)
        else:
            ciclo_end = date(hoje.year, hoje.month + 1, 18)
    else:
        # previous month start
        if hoje.month == 1:
            ciclo_start = date(hoje.year - 1, 12, 19)
        else:
            ciclo_start = date(hoje.year, hoje.month - 1, 19)
        ciclo_end = date(hoje.year, hoje.month, 18)

    # Use ciclo_start (00:00) to ciclo_end (23:59:59)
    from datetime import time
    dt_start = datetime.combine(ciclo_start, time.min)
    dt_end = datetime.combine(ciclo_end, time.max)

    qs = Linha.objects.filter(criado_em__gte=dt_start, criado_em__lte=dt_end)

    total_linhas = qs.count()
    linhas_ativas = qs.filter(ativa=True).count()
    linhas_inativas = qs.filter(ativa=False).count()
    media_valor_plano = qs.aggregate(avg_valor=Avg('valor_plano'))['avg_valor'] or 0

    # Novas linhas no período do ciclo: only from ESTOQUE (movimentadas) and PORTABILIDADE
    action_values = ['ESTOQUE', 'PORTABILIDADE']
    action_qs = qs.filter(acao__in=action_values)
    novas_30 = action_qs.count()

    # Distribuição por tipo de plano (top 10)
    tipo_plano_qs = qs.values('tipo_plano').annotate(count=Count('id')).order_by('-count')
    # Top 5, aggregate the rest as 'Outros'
    tipo_plano_list = list(tipo_plano_qs)
    top = tipo_plano_list[:5]
    outros = tipo_plano_list[5:]
    outros_count = sum([t['count'] for t in outros])
    tipo_plano_labels = [t['tipo_plano'] for t in top]
    tipo_plano_data = [t['count'] for t in top]
    if outros_count:
        tipo_plano_labels.append('Outros')
        tipo_plano_data.append(outros_count)

    # Novas linhas por dia (últimos 30 dias)
    dias = []
    contagens_dias = []
    # Series per day for the cycle period (from ciclo_start to ciclo_end)
    total_days = (ciclo_end - ciclo_start).days + 1
    for i in range(total_days):
        dia = ciclo_start + timedelta(days=i)
        # count only actions ESTOQUE or PORTABILIDADE on that day
        dia_qs = Linha.objects.filter(criado_em__date=dia, acao__in=action_values)
        count = dia_qs.count()
        dias.append(dia.strftime('%d/%m'))
        contagens_dias.append(count)

    # Protocol counts per day by status for the ciclo period
    protocolo_statuses = ['pendente', 'resolvido', 'cancelado']
    protocolos_por_status = {s: [] for s in protocolo_statuses}
    for i in range(total_days):
        dia = ciclo_start + timedelta(days=i)
        for s in protocolo_statuses:
            c = Protocolo.objects.filter(status=s, criado_em__date=dia).count()
            protocolos_por_status[s].append(c)

    # Pending protocols
    protocolos_pendentes_qs = Protocolo.objects.filter(status='pendente').order_by('-criado_em')
    protocolos_pendentes_count = protocolos_pendentes_qs.count()
    
    # Completed protocols (resolvido status)
    protocolos_concluidos_qs = Protocolo.objects.filter(status='resolvido').order_by('-criado_em')
    protocolos_concluidos_count = protocolos_concluidos_qs.count()

    import json

    # format average as float or zero
    try:
        media_valor_num = float(media_valor_plano)
    except Exception:
        media_valor_num = 0.0

    context = {
        'total_linhas': total_linhas,
        'linhas_ativas': linhas_ativas,
        'linhas_inativas': linhas_inativas,
        # explicit alias for cancelled lines (business wants "canceladas")
        'linhas_canceladas': linhas_inativas,
        'novas_30': novas_30,
        'media_valor_plano': f"R$ {media_valor_num:,.2f}",
    'ciclo_start': ciclo_start.strftime('%Y-%m-%d'),
    'ciclo_end': ciclo_end.strftime('%Y-%m-%d'),
        # also provide JSON-serialized strings to safely embed in JS without json_script
        'tipo_plano_labels_json': json.dumps(tipo_plano_labels, ensure_ascii=False),
        'tipo_plano_data_json': json.dumps(tipo_plano_data),
        'dias_json': json.dumps(dias, ensure_ascii=False),
        'contagens_dias_json': json.dumps(contagens_dias),
    'protocolos_pendentes_count': protocolos_pendentes_count,
    'protocolos_concluidos_count': protocolos_concluidos_count,
    # protocol time series for charts
    'protocolos_dias_labels_json': json.dumps(dias, ensure_ascii=False),
    'protocolos_pendente_series_json': json.dumps(protocolos_por_status['pendente']),
    'protocolos_resolvido_series_json': json.dumps(protocolos_por_status['resolvido']),
    'protocolos_cancelado_series_json': json.dumps(protocolos_por_status['cancelado']),
    }
    return render(request, 'dashboard.html', context)


@login_required
def protocolo(request):
    """
    Página de protocolo que mostra todos os pedidos criados no menu Nova linha e Fidelidade
    """
    from django.db.models import Q
    from django.core.paginator import Paginator
    from datetime import datetime, timedelta
    from django.contrib.auth.models import User
    from .models import Fidelidade
    
    # Buscar parâmetros de filtro
    filtro_tipo = request.GET.get('tipo', 'todos')  # todos, linhas, fidelidades
    busca = request.GET.get('busca', '').strip()
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    usuario_criador = request.GET.get('usuario', '')
    
    # Preparar querysets
    linhas_qs = Linha.objects.select_related('criado_por', 'cliente').all()
    fidelidades_qs = Fidelidade.objects.select_related('criado_por', 'linha').all()
    
    # Aplicar filtros de data
    if data_inicio:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            linhas_qs = linhas_qs.filter(criado_em__gte=data_inicio_dt)
            fidelidades_qs = fidelidades_qs.filter(criado_em__gte=data_inicio_dt)
        except ValueError:
            pass
    
    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d')
            # Incluir todo o dia
            data_fim_dt = data_fim_dt.replace(hour=23, minute=59, second=59)
            linhas_qs = linhas_qs.filter(criado_em__lte=data_fim_dt)
            fidelidades_qs = fidelidades_qs.filter(criado_em__lte=data_fim_dt)
        except ValueError:
            pass
    
    # Filtro por usuário criador
    if usuario_criador:
        linhas_qs = linhas_qs.filter(criado_por__username__icontains=usuario_criador)
        fidelidades_qs = fidelidades_qs.filter(criado_por__username__icontains=usuario_criador)
    
    # Filtro de busca (número, empresa, observações)
    if busca:
        linhas_qs = linhas_qs.filter(
            Q(numero__icontains=busca) |
            Q(empresa__icontains=busca) |
            Q(cnpj__icontains=busca) |
            Q(rp__icontains=busca) |
            Q(observacoes__icontains=busca)
        )
        fidelidades_qs = fidelidades_qs.filter(
            Q(linha__numero__icontains=busca) |
            Q(linha__empresa__icontains=busca) |
            Q(observacoes__icontains=busca)
        )
    
    # Aplicar filtro de tipo
    requests_list = []
    
    if filtro_tipo in ['todos', 'linhas']:
        # Adicionar linhas à lista de requests
        for linha in linhas_qs.order_by('-criado_em'):
            requests_list.append({
                'tipo': 'Nova Linha',
                'id': linha.id,
                'numero': linha.numero,
                'cliente': linha.empresa or 'N/A',
                'cnpj': linha.cnpj or 'N/A',
                'rp': linha.rp or 'N/A',
                'tipo_plano': linha.tipo_plano,
                'valor': linha.valor_plano,
                'acao': linha.acao,
                'status': 'Ativa' if linha.ativa else 'Inativa',
                'observacoes': linha.observacoes or 'Sem observações',
                'criado_por': linha.criado_por.username if linha.criado_por else 'Sistema',
                'criado_em': linha.criado_em,
                'detalhes_extra': {
                    'iccid': linha.iccid or 'N/A',
                    'taxa_manutencao': linha.taxa_manutencao or 0,
                }
            })
    
    if filtro_tipo in ['todos', 'fidelidades']:
        # Adicionar fidelidades à lista de requests
        for fidelidade in fidelidades_qs.order_by('-criado_em'):
            requests_list.append({
                'tipo': 'Fidelidade',
                'id': fidelidade.id,
                'numero': fidelidade.linha.numero,
                'cliente': fidelidade.linha.empresa or 'N/A',
                'cnpj': fidelidade.linha.cnpj or 'N/A',
                'rp': fidelidade.linha.rp or 'N/A',
                'tipo_plano': fidelidade.linha.tipo_plano,
                'valor': fidelidade.linha.valor_plano,
                'acao': 'FIDELIDADE',
                'status': 'Processada',
                'observacoes': fidelidade.observacoes,
                'criado_por': fidelidade.criado_por.username if fidelidade.criado_por else 'Sistema',
                'criado_em': fidelidade.criado_em,
                'detalhes_extra': {
                    'linha_ativa': 'Sim' if fidelidade.linha.ativa else 'Não',
                    'tipo_original': 'Registro de Fidelidade',
                }
            })
    
    # Ordenar por data de criação (mais recente primeiro)
    requests_list.sort(key=lambda x: x['criado_em'], reverse=True)
    
    # Paginação
    paginator = Paginator(requests_list, 15)  # 15 itens por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_linhas = Linha.objects.count()
    total_fidelidades = Fidelidade.objects.count()
    total_requests = total_linhas + total_fidelidades
    
    # Requests do último mês
    ultimo_mes = timezone.now() - timedelta(days=30)
    linhas_ultimo_mes = Linha.objects.filter(criado_em__gte=ultimo_mes).count()
    fidelidades_ultimo_mes = Fidelidade.objects.filter(criado_em__gte=ultimo_mes).count()
    
    # Lista de usuários para filtro
    usuarios = User.objects.filter(
        Q(linha__isnull=False) | Q(fidelidade__isnull=False)
    ).distinct().values('username').order_by('username')
    
    context = {
        'page_obj': page_obj,
        'filtro_tipo': filtro_tipo,
        'busca': busca,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'usuario_criador': usuario_criador,
        'total_requests': total_requests,
        'total_linhas': total_linhas,
        'total_fidelidades': total_fidelidades,
        'linhas_ultimo_mes': linhas_ultimo_mes,
        'fidelidades_ultimo_mes': fidelidades_ultimo_mes,
        'usuarios': usuarios,
        'hoje': timezone.now().date(),
    }
    
    return render(request, 'protocolo.html', context)


@login_required
def relatorios(request):
    # Provide distinct tipo_plano values so template can render a select
    tipo_planos = Linha.objects.order_by('tipo_plano').values_list('tipo_plano', flat=True).distinct()
    context = {
        'tipo_plano_choices': list(tipo_planos),
    }
    return render(request, 'relatorios.html', context)


@login_required
def configuracoes(request):
    # Lista simples de usuários atualmente bloqueados (se houver)
    bloqueados = LoginAttempt.objects.filter(is_blocked=True, blocked_until__gt=timezone.now()) \
        .values_list('username', flat=True).distinct()
    return render(request, 'configuracoes.html', { 'bloqueados': list(bloqueados) })

@login_required
def desbloquear_usuario(request):
    """Desbloqueia manualmente um usuário removendo registros de LoginAttempt bloqueados.

    Regras:
    - Apenas usuários staff podem executar.
    - Recebe POST com campo 'username'.
    - Remove tentativas para aquele username e IP (todas) marcadas como bloqueadas.
    - Exibe mensagem de sucesso ou erro e redireciona para configurações.
    """
    if not request.user.is_staff:
        messages.error(request, 'Você não tem permissão para desbloquear usuários.')
        return redirect('linhas:configuracoes')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        if not username:
            messages.error(request, 'Informe o nome de usuário para desbloquear.')
            return redirect('linhas:configuracoes')

        # Remover todas as tentativas para o username (independente do IP) para reset total
        attempts = LoginAttempt.objects.filter(username=username)
        count = attempts.count()
        if count == 0:
            messages.warning(request, f'Nenhum bloqueio ativo encontrado para "{username}".')
        else:
            attempts.delete()
            messages.success(request, f'Usuário "{username}" desbloqueado. ({count} registro(s) removido(s))')
        return redirect('linhas:configuracoes')

    return redirect('linhas:configuracoes')


@login_required
def export_linhas_cycle_csv(request):
    # compute ciclo as in dashboard
    from django.utils import timezone
    from datetime import datetime, date, time
    hoje = timezone.now().date()
    if hoje.day >= 19:
        ciclo_start = date(hoje.year, hoje.month, 19)
        if hoje.month == 12:
            ciclo_end = date(hoje.year + 1, 1, 18)
        else:
            ciclo_end = date(hoje.year, hoje.month + 1, 18)
    else:
        if hoje.month == 1:
            ciclo_start = date(hoje.year - 1, 12, 19)
        else:
            ciclo_start = date(hoje.year, hoje.month - 1, 19)
        ciclo_end = date(hoje.year, hoje.month, 18)

    # default period is the billing cycle
    dt_start = datetime.combine(ciclo_start, time.min)
    dt_end = datetime.combine(ciclo_end, time.max)

    # Allow overrides from GET params: periodo_start / periodo_end (YYYY-MM-DD)
    periodo_start = request.GET.get('periodo_start')
    periodo_end = request.GET.get('periodo_end')
    from datetime import datetime as _dt
    try:
        if periodo_start:
            dt_start = _dt.combine(_dt.strptime(periodo_start, '%Y-%m-%d').date(), time.min)
        if periodo_end:
            dt_end = _dt.combine(_dt.strptime(periodo_end, '%Y-%m-%d').date(), time.max)
        # server-side validation: start must not be after end
        if periodo_start and periodo_end and dt_start > dt_end:
            messages.error(request, 'Período inválido: data de início posterior à data fim.')
            return redirect('linhas:relatorios')
    except Exception:
        # ignore parse errors and keep the cycle defaults
        pass

    qs = Linha.objects.filter(criado_em__gte=dt_start, criado_em__lte=dt_end)

    # Apply optional filters: empresa, cnpj, tipo_plano, numero, status
    empresa = request.GET.get('empresa')
    cnpj = request.GET.get('cnpj')
    tipo_plano = request.GET.get('tipo_plano')
    numero = request.GET.get('numero')
    status = request.GET.get('status')

    # server-side CNPJ validation (if provided)
    if cnpj:
        import re
        cnpj_digits = re.sub(r'\D', '', cnpj)
        if len(cnpj_digits) != 14:
            messages.error(request, 'CNPJ inválido: verifique o número informado.')
            return redirect('linhas:relatorios')
        # normalize to the stored format by filtering with contains on digits
        qs = qs.filter(cnpj__icontains=cnpj_digits)

    if empresa:
        qs = qs.filter(empresa__icontains=empresa)
    if tipo_plano:
        qs = qs.filter(tipo_plano__icontains=tipo_plano)

    # Behavior requirement: when a CNPJ or empresa is provided, numero filter should only
    # be applied if the given numero actually belongs to that CNPJ/empresa. Otherwise
    # return all lines linked to that CNPJ/empresa (ignore numero filter).
    if numero:
        if cnpj or empresa:
            # check within the already filtered qs (which contains only lines for that cnpj/empresa)
            if qs.filter(numero__icontains=numero).exists():
                qs = qs.filter(numero__icontains=numero)
            else:
                # numero not associated with this CNPJ/empresa — ignore numero filter
                pass
        else:
            # no cnpj/empresa context — apply numero filter globally
            qs = qs.filter(numero__icontains=numero)

    if status:
        # map status to active flag
        if status == 'ativa':
            qs = qs.filter(ativa=True)
        elif status == 'cancelada':
            qs = qs.filter(ativa=False)

    # CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="linhas_{ciclo_start}_{ciclo_end}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Numero', 'ICCID', 'Empresa', 'CNPJ', 'RP', 'Tipo Plano', 'Valor Plano', 'Ação', 'Ativa', 'Criado Em'])
    for l in qs:
        writer.writerow([l.numero, l.iccid, l.empresa, l.cnpj, l.rp, l.tipo_plano, str(l.valor_plano), l.acao, l.ativa, l.criado_em.strftime('%Y-%m-%d %H:%M')])
    return response


@login_required
def export_linhas_cycle_xlsx(request):
    """Generate an XLSX file with the same filters as CSV export."""
    # compute ciclo as in dashboard
    from django.utils import timezone
    from datetime import datetime, date, time
    hoje = timezone.now().date()
    if hoje.day >= 19:
        ciclo_start = date(hoje.year, hoje.month, 19)
        if hoje.month == 12:
            ciclo_end = date(hoje.year + 1, 1, 18)
        else:
            ciclo_end = date(hoje.year, hoje.month + 1, 18)
    else:
        if hoje.month == 1:
            ciclo_start = date(hoje.year - 1, 12, 19)
        else:
            ciclo_start = date(hoje.year, hoje.month - 1, 19)
        ciclo_end = date(hoje.year, hoje.month, 18)

    dt_start = datetime.combine(ciclo_start, time.min)
    dt_end = datetime.combine(ciclo_end, time.max)

    # Allow overrides from GET params: periodo_start / periodo_end (YYYY-MM-DD)
    periodo_start = request.GET.get('periodo_start')
    periodo_end = request.GET.get('periodo_end')
    from datetime import datetime as _dt
    try:
        if periodo_start:
            dt_start = _dt.combine(_dt.strptime(periodo_start, '%Y-%m-%d').date(), time.min)
        if periodo_end:
            dt_end = _dt.combine(_dt.strptime(periodo_end, '%Y-%m-%d').date(), time.max)
        if periodo_start and periodo_end and dt_start > dt_end:
            messages.error(request, 'Período inválido: data de início posterior à data fim.')
            return redirect('linhas:relatorios')
    except Exception:
        pass

    qs = Linha.objects.filter(criado_em__gte=dt_start, criado_em__lte=dt_end)

    # same optional filters as CSV
    empresa = request.GET.get('empresa')
    cnpj = request.GET.get('cnpj')
    tipo_plano = request.GET.get('tipo_plano')
    numero = request.GET.get('numero')
    status = request.GET.get('status')

    if cnpj:
        import re
        cnpj_digits = re.sub(r'\D', '', cnpj)
        if len(cnpj_digits) != 14:
            messages.error(request, 'CNPJ inválido: verifique o número informado.')
            return redirect('linhas:relatorios')
        qs = qs.filter(cnpj__icontains=cnpj_digits)

    if empresa:
        qs = qs.filter(empresa__icontains=empresa)
    if tipo_plano:
        qs = qs.filter(tipo_plano__icontains=tipo_plano)

    # Same business rule as CSV: if CNPJ or empresa provided, only apply numero filter
    # when the numero exists for that CNPJ/empresa; otherwise ignore numero and return
    # all lines for that CNPJ/empresa.
    if numero:
        if cnpj or empresa:
            if qs.filter(numero__icontains=numero).exists():
                qs = qs.filter(numero__icontains=numero)
            else:
                pass
        else:
            qs = qs.filter(numero__icontains=numero)

    if status:
        if status == 'ativa':
            qs = qs.filter(ativa=True)
        elif status == 'cancelada':
            qs = qs.filter(ativa=False)

    # generate xlsx in-memory
    from io import BytesIO
    try:
        from openpyxl import Workbook
    except Exception:
        # openpyxl not installed; return a helpful message
        messages.error(request, 'openpyxl não está instalado no servidor. Instale via pip install openpyxl')
        return redirect('linhas:relatorios')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Linhas'
    headers = ['Numero', 'ICCID', 'Empresa', 'CNPJ', 'RP', 'Tipo Plano', 'Valor Plano', 'Ação', 'Ativa', 'Criado Em']
    ws.append(headers)
    for l in qs:
        ws.append([l.numero, l.iccid, l.empresa, l.cnpj, l.rp, l.tipo_plano, float(l.valor_plano or 0), l.acao, 'Sim' if l.ativa else 'Não', l.criado_em.strftime('%Y-%m-%d %H:%M')])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="linhas_{ciclo_start}_{ciclo_end}.xlsx"'
    return response


@login_required
def export_protocolos_pendentes_csv(request):
    # Allow optional date filters when exporting protocolos
    periodo_start = request.GET.get('periodo_start')
    periodo_end = request.GET.get('periodo_end')
    qs = Protocolo.objects.filter(status='pendente')
    from datetime import datetime
    try:
        if periodo_start:
            ds = datetime.strptime(periodo_start, '%Y-%m-%d')
            qs = qs.filter(criado_em__gte=ds)
        if periodo_end:
            de = datetime.strptime(periodo_end, '%Y-%m-%d')
            # include end of day
            from datetime import time as _time
            de = datetime.combine(de.date(), _time.max)
            qs = qs.filter(criado_em__lte=de)
        # server-side validation: if both provided, ensure start <= end
        if periodo_start and periodo_end:
            # compare by date
            ds_check = datetime.strptime(periodo_start, '%Y-%m-%d')
            de_check = datetime.strptime(periodo_end, '%Y-%m-%d')
            if ds_check > de_check:
                messages.error(request, 'Período inválido: data de início posterior à data fim.')
                return redirect('linhas:relatorios')
    except Exception:
        # ignore parse errors
        pass
    qs = qs.order_by('-criado_em')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="protocolos_pendentes.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID', 'Titulo', 'Descricao', 'Criado Por', 'Criado Em', 'Status'])
    for p in qs:
        writer.writerow([p.id, p.titulo, (p.descricao or '')[:200], (p.criado_por.username if p.criado_por else ''), p.criado_em.strftime('%Y-%m-%d %H:%M'), p.status])
    return response

@login_required
def fidelidade(request):
    """Página limpa para registro simples de fidelidade: número da linha + observações."""
    if request.method == 'POST':
        from .models import Linha, Fidelidade
        numero = request.POST.get('numero_linha', '').strip()
        observacoes = request.POST.get('observacoes', '').strip()

        if not numero:
            messages.error(request, 'Informe o número da linha.')
        elif not observacoes:
            messages.error(request, 'Informe as observações.')
        elif len(observacoes) < 5:
            messages.error(request, 'Observações devem ter pelo menos 5 caracteres.')
        else:
            try:
                linha = Linha.objects.get(numero=numero)
            except Linha.DoesNotExist:
                messages.error(request, f'Linha {numero} não encontrada.')
            else:
                Fidelidade.objects.create(linha=linha, observacoes=observacoes, criado_por=request.user)
                messages.success(request, f'Fidelidade registrada para a linha {numero}.')
                return redirect('linhas:fidelidade')

    return render(request, 'linhas/fidelidade.html')

def buscar_linha_dados(request):
    """
    AJAX endpoint para buscar dados da linha (cliente e RP)
    baseado no número da linha selecionado
    """
    numero = request.GET.get('numero', '')
    
    if not numero:
        return JsonResponse({
            'success': False,
            'error': 'Número da linha não informado'
        })
    
    try:
        linha = Linha.objects.filter(numero=numero).first()
        
        if linha:
            # Buscar nome do cliente baseado no CNPJ da linha
            cliente_nome = ""
            if linha.cliente:
                cliente_nome = linha.cliente.empresa
            elif linha.empresa:
                cliente_nome = linha.empresa
            
            return JsonResponse({
                'success': True,
                'dados': {
                    'cliente': cliente_nome,
                    'rp': linha.rp or '',
                    'empresa': linha.empresa or '',
                    'cnpj': linha.cnpj or '',
                    'ativa': linha.ativa,
                    'tipo_plano': linha.tipo_plano or '',
                    'operadora': linha.operadora or '',
                    'valor_plano': str(linha.valor_plano) if linha.valor_plano else '',
                    'iccid': linha.iccid or '',
                    'status': 'Ativa' if linha.ativa else 'Inativa'
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Linha não encontrada'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro ao buscar dados da linha: {str(e)}'
        })

@login_required
def apagar_fidelidade(request):
    """
    AJAX endpoint para apagar todas as fidelidades de uma linha específica
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método não permitido'
        })
    
    numero = request.GET.get('numero', '')
    
    if not numero:
        return JsonResponse({
            'success': False,
            'error': 'Número da linha não informado'
        })
    
    try:
        from .models import Linha, Fidelidade
        
        # Buscar a linha
        linha = Linha.objects.filter(numero=numero).first()
        
        if not linha:
            return JsonResponse({
                'success': False,
                'error': f'Linha {numero} não encontrada'
            })
        
        # Buscar e contar fidelidades da linha
        fidelidades = Fidelidade.objects.filter(linha=linha)
        count = fidelidades.count()
        
        if count == 0:
            return JsonResponse({
                'success': False,
                'error': f'Nenhuma fidelidade encontrada para a linha {numero}'
            })
        
        # Apagar todas as fidelidades da linha
        fidelidades.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'{count} fidelidade(s) apagada(s) com sucesso para a linha {numero}'
        })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro ao apagar fidelidade: {str(e)}'
        })

@login_required
def criar_usuario_empresa(request):
    """
    View para criar usuário empresa com dados completos
    """
    from django.contrib.auth.models import User
    from django.db import transaction
    from django.http import JsonResponse
    import re
    
    if request.method == 'POST':
        
        try:
            # Importar o modelo dentro do try para capturar possíveis erros
            from .models import UsuarioEmpresa
            
            with transaction.atomic():
                # Coletar dados do formulário
                cnpj = request.POST.get('cnpj', '').strip()
                razao_social = request.POST.get('razao_social', '').strip()
                nome_fantasia = request.POST.get('nome_fantasia', '').strip()
                endereco = request.POST.get('endereco', '').strip()
                email = request.POST.get('email', '').strip()
                telefone = request.POST.get('telefone', '').strip()
                nome_completo_agente = request.POST.get('nome_completo_agente', '').strip()
                cpf_responsavel = request.POST.get('cpf_responsavel', '').strip()
                data_nascimento = request.POST.get('data_nascimento', '').strip()
                username = request.POST.get('username', '').strip()
                senha = request.POST.get('senha', '').strip()
                is_administrador = request.POST.get('is_administrador') == 'on'
                
                # Validações básicas
                if not all([cnpj, razao_social, endereco, email, telefone, nome_completo_agente, cpf_responsavel, data_nascimento, username, senha]):
                    return JsonResponse({
                        'success': False,
                        'error': 'Todos os campos obrigatórios devem ser preenchidos.'
                    })
                
                # Validar se usuário já existe
                if User.objects.filter(username=username).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'Nome de usuário "{username}" já existe. Escolha outro.'
                    })
                
                # Validar se email já existe
                if User.objects.filter(email=email).exists():
                    return JsonResponse({
                        'success': False,
                        'error': f'Email "{email}" já está em uso. Use outro email.'
                    })
                
                # Validar CNPJ (básico - apenas dígitos)
                cnpj_digits = re.sub(r'\D', '', cnpj)
                if len(cnpj_digits) != 14:
                    return JsonResponse({
                        'success': False,
                        'error': 'CNPJ deve ter 14 dígitos válidos.'
                    })
                
                # Validar CPF (básico - apenas dígitos)
                cpf_digits = re.sub(r'\D', '', cpf_responsavel)
                if len(cpf_digits) != 11:
                    return JsonResponse({
                        'success': False,
                        'error': 'CPF deve ter 11 dígitos válidos.'
                    })
                
                # Criar usuário Django
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=senha,
                    first_name=razao_social[:30],  # Limita para o campo first_name
                    is_staff=is_administrador,
                    is_superuser=is_administrador
                )
                
                # Criar perfil empresarial
                usuario_empresa = UsuarioEmpresa.objects.create(
                    user=user,
                    cnpj=cnpj,
                    razao_social=razao_social,
                    nome_fantasia=nome_fantasia,
                    endereco=endereco,
                    telefone=telefone,
                    nome_completo_agente=nome_completo_agente,
                    cpf_agente=cpf_responsavel,
                    data_nascimento_agente=data_nascimento,
                    criado_por=request.user
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Usuário "{razao_social}" ({username}) criado com sucesso!',
                    'usuario': {
                        'id': user.id,
                        'username': username,
                        'razao_social': razao_social,
                        'cnpj': cnpj,
                        'is_administrador': is_administrador
                    }
                })
                
        except ImportError as e:
            return JsonResponse({
                'success': False,
                'error': 'Erro de configuração do sistema. Contate o administrador.'
            })
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Erro detalhado: {error_details}")  # Para debug no console
            return JsonResponse({
                'success': False,
                'error': f'Erro ao criar usuário: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método não permitido'
    })


@login_required
def listar_usuarios_empresa(request):
    """
    View para listar todos os usuários empresa cadastrados
    """
    from django.core.paginator import Paginator
    from django.db.models import Q
    from .models import UsuarioEmpresa
    
    # Buscar todos os usuários empresa
    usuarios_qs = UsuarioEmpresa.objects.select_related('user', 'criado_por').all()
    
    # Filtros opcionais
    search = request.GET.get('search', '').strip()
    if search:
        usuarios_qs = usuarios_qs.filter(
            Q(razao_social__icontains=search) |
            Q(nome_fantasia__icontains=search) |
            Q(cnpj__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    # Filtro por status
    status = request.GET.get('status', '')
    if status == 'ativo':
        usuarios_qs = usuarios_qs.filter(user__is_active=True)
    elif status == 'inativo':
        usuarios_qs = usuarios_qs.filter(user__is_active=False)
    elif status == 'admin':
        usuarios_qs = usuarios_qs.filter(user__is_staff=True)
    
    # Ordenação
    usuarios_qs = usuarios_qs.order_by('-criado_em')
    
    # Paginação
    paginator = Paginator(usuarios_qs, 10)  # 10 usuários por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_usuarios = UsuarioEmpresa.objects.count()
    usuarios_ativos = UsuarioEmpresa.objects.filter(user__is_active=True).count()
    usuarios_admin = UsuarioEmpresa.objects.filter(user__is_staff=True).count()
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'status_filter': status,
        'total_usuarios': total_usuarios,
        'usuarios_ativos': usuarios_ativos,
        'usuarios_admin': usuarios_admin,
        'usuarios_inativos': total_usuarios - usuarios_ativos,
    }
    
    return render(request, 'linhas/listar_usuarios.html', context)


@login_required
def toggle_usuario_status(request):
    """
    AJAX view para ativar/desativar usuário
    """
    if request.method == 'POST':
        from django.contrib.auth.models import User
        
        user_id = request.POST.get('user_id')
        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'ID do usuário não informado'
            })
        
        try:
            user = User.objects.get(id=user_id)
            user.is_active = not user.is_active
            user.save()
            
            status_text = 'ativado' if user.is_active else 'desativado'
            
            return JsonResponse({
                'success': True,
                'message': f'Usuário "{user.username}" {status_text} com sucesso!',
                'new_status': user.is_active
            })
            
        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuário não encontrado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erro ao alterar status: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método não permitido'
    })


@login_required
def visualizar_usuario(request, user_id):
    """
    AJAX view para buscar detalhes completos do usuário
    """
    try:
        from .models import UsuarioEmpresa
        usuario_empresa = UsuarioEmpresa.objects.select_related('user', 'criado_por').get(user__id=user_id)
        
        dados = {
            'id': usuario_empresa.id,
            'username': usuario_empresa.user.username,
            'email': usuario_empresa.user.email,
            'first_name': usuario_empresa.user.first_name,
            'is_active': usuario_empresa.user.is_active,
            'is_staff': usuario_empresa.user.is_staff,
            'date_joined': usuario_empresa.user.date_joined.strftime('%d/%m/%Y %H:%M'),
            'last_login': usuario_empresa.user.last_login.strftime('%d/%m/%Y %H:%M') if usuario_empresa.user.last_login else 'Nunca',
            'cnpj': usuario_empresa.cnpj,
            'razao_social': usuario_empresa.razao_social,
            'nome_fantasia': usuario_empresa.nome_fantasia,
            'endereco': usuario_empresa.endereco,
            'telefone': usuario_empresa.telefone,
            'nome_completo_agente': usuario_empresa.nome_completo_agente,
            'cpf_agente': usuario_empresa.cpf_agente,
            'data_nascimento_agente': usuario_empresa.data_nascimento_agente.strftime('%d/%m/%Y'),
            'criado_em': usuario_empresa.criado_em.strftime('%d/%m/%Y %H:%M'),
            'criado_por': usuario_empresa.criado_por.username if usuario_empresa.criado_por else 'Sistema',
            'atualizado_em': usuario_empresa.atualizado_em.strftime('%d/%m/%Y %H:%M'),
        }
        
        return JsonResponse({
            'success': True,
            'usuario': dados
        })
        
    except UsuarioEmpresa.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Usuário não encontrado'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro ao buscar dados: {str(e)}'
        })


@login_required
def editar_usuario(request, user_id):
    """
    View para editar dados do usuário empresa
    """
    from .models import UsuarioEmpresa
    from django.contrib.auth.models import User
    from django.shortcuts import get_object_or_404
    from django.db import transaction
    
    usuario_empresa = get_object_or_404(UsuarioEmpresa, user__id=user_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Atualizar dados do User
                user = usuario_empresa.user
                user.email = request.POST.get('email', '').strip()
                user.first_name = request.POST.get('razao_social', '')[:30]
                user.is_staff = request.POST.get('is_administrador') == 'on'
                user.is_superuser = user.is_staff
                
                # Verificar se email não está em uso por outro usuário
                if User.objects.filter(email=user.email).exclude(id=user.id).exists():
                    messages.error(request, f'Email "{user.email}" já está em uso por outro usuário.')
                    return render(request, 'linhas/editar_usuario.html', {'usuario_empresa': usuario_empresa})
                
                user.save()
                
                # Atualizar dados da UsuarioEmpresa
                usuario_empresa.cnpj = request.POST.get('cnpj', '').strip()
                usuario_empresa.razao_social = request.POST.get('razao_social', '').strip()
                usuario_empresa.nome_fantasia = request.POST.get('nome_fantasia', '').strip()
                usuario_empresa.endereco = request.POST.get('endereco', '').strip()
                usuario_empresa.telefone = request.POST.get('telefone', '').strip()
                usuario_empresa.nome_completo_agente = request.POST.get('nome_completo_agente', '').strip()
                usuario_empresa.cpf_agente = request.POST.get('cpf_agente', '').strip()
                usuario_empresa.data_nascimento_agente = request.POST.get('data_nascimento_agente', '')
                
                usuario_empresa.save()
                
                messages.success(request, f'Usuário "{usuario_empresa.razao_social}" atualizado com sucesso!')
                return redirect('linhas:listar_usuarios_empresa')
                
        except Exception as e:
            messages.error(request, f'Erro ao atualizar usuário: {str(e)}')
    
    return render(request, 'linhas/editar_usuario.html', {'usuario_empresa': usuario_empresa})


@login_required
def excluir_usuario(request, user_id):
    """
    View para excluir usuário empresa
    """
    from .models import UsuarioEmpresa
    from django.shortcuts import get_object_or_404
    
    usuario_empresa = get_object_or_404(UsuarioEmpresa, user__id=user_id)
    
    if request.method == 'POST':
        try:
            razao_social = usuario_empresa.razao_social
            username = usuario_empresa.user.username
            
            # Verificar se não é o próprio usuário logado
            if usuario_empresa.user == request.user:
                messages.error(request, 'Você não pode excluir sua própria conta.')
                return redirect('linhas:listar_usuarios_empresa')
            
            # Excluir usuário (cascata excluirá UsuarioEmpresa)
            usuario_empresa.user.delete()
            
            messages.success(request, f'Usuário "{razao_social}" ({username}) excluído com sucesso!')
            return redirect('linhas:listar_usuarios_empresa')
            
        except Exception as e:
            messages.error(request, f'Erro ao excluir usuário: {str(e)}')
    
    return render(request, 'linhas/excluir_usuario.html', {'usuario_empresa': usuario_empresa})


@login_required
def alterar_senhas(request):
    """
    View para alterar senhas de usuários empresa
    """
    from .models import UsuarioEmpresa
    from django.contrib.auth.models import User
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Buscar todos os usuários empresa para alterar senha
    usuarios_qs = UsuarioEmpresa.objects.select_related('user').filter(user__is_active=True).order_by('razao_social')
    
    # Filtro de busca
    search = request.GET.get('search', '').strip()
    if search:
        usuarios_qs = usuarios_qs.filter(
            Q(razao_social__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search)
        )
    
    # Paginação
    paginator = Paginator(usuarios_qs, 15)  # 15 usuários por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'total_usuarios': usuarios_qs.count(),
    }
    
    return render(request, 'linhas/alterar_senhas.html', context)


@login_required
def alterar_senha_usuario(request):
    """
    AJAX view para alterar senha de um usuário específico
    """
    if request.method == 'POST':
        from django.contrib.auth.models import User
        import re
        
        user_id = request.POST.get('user_id')
        nova_senha = request.POST.get('nova_senha', '').strip()
        confirmar_senha = request.POST.get('confirmar_senha', '').strip()
        
        # Validações
        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'ID do usuário não informado'
            })
        
        if not nova_senha or not confirmar_senha:
            return JsonResponse({
                'success': False,
                'error': 'Nova senha e confirmação são obrigatórias'
            })
        
        if nova_senha != confirmar_senha:
            return JsonResponse({
                'success': False,
                'error': 'Nova senha e confirmação não coincidem'
            })
        
        if len(nova_senha) < 6:
            return JsonResponse({
                'success': False,
                'error': 'A senha deve ter pelo menos 6 caracteres'
            })
        
        # Validar complexidade da senha
        if not re.search(r'[A-Za-z]', nova_senha) or not re.search(r'\d', nova_senha):
            return JsonResponse({
                'success': False,
                'error': 'A senha deve conter pelo menos uma letra e um número'
            })
        
        try:
            user = User.objects.get(id=user_id)
            
            # Não permitir alterar senha do próprio usuário por segurança
            if user == request.user:
                return JsonResponse({
                    'success': False,
                    'error': 'Você não pode alterar sua própria senha por esta interface. Use as configurações da conta.'
                })
            
            # Alterar a senha
            user.set_password(nova_senha)
            user.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Senha do usuário "{user.username}" alterada com sucesso!'
            })
            
        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuário não encontrado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erro ao alterar senha: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método não permitido'
    })


@login_required
def cadastro_empresa(request):
    """
    View para cadastro de empresa com dados completos
    """
    from django.db import transaction
    import re
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Coletar dados do formulário
                cnpj = request.POST.get('cnpj', '').strip()
                razao_social = request.POST.get('razao_social', '').strip()
                nome_fantasia = request.POST.get('nome_fantasia', '').strip()
                nome_proprietario = request.POST.get('nome_proprietario', '').strip()
                cpf_proprietario = request.POST.get('cpf_proprietario', '').strip()
                data_nascimento = request.POST.get('data_nascimento', '').strip()
                endereco = request.POST.get('endereco', '').strip()
                telefone = request.POST.get('telefone', '').strip()
                email = request.POST.get('email', '').strip()
                arquivo_linhas = request.FILES.get('arquivo_linhas')
                
                # Validações básicas
                if not all([cnpj, razao_social, nome_proprietario, cpf_proprietario, data_nascimento, endereco, telefone, email]):
                    messages.error(request, 'Todos os campos obrigatórios devem ser preenchidos.')
                    return render(request, 'linhas/cadastro_empresa.html')
                
                # Validar CNPJ (básico - apenas dígitos)
                cnpj_digits = re.sub(r'\D', '', cnpj)
                if len(cnpj_digits) != 14:
                    messages.error(request, 'CNPJ deve ter 14 dígitos válidos.')
                    return render(request, 'linhas/cadastro_empresa.html')
                
                # Verificar se CNPJ já existe
                if Cliente.objects.filter(cnpj=cnpj).exists():
                    messages.error(request, f'CNPJ {cnpj} já está cadastrado no sistema.')
                    return render(request, 'linhas/cadastro_empresa.html')
                
                # Validar CPF (básico - apenas dígitos)
                cpf_digits = re.sub(r'\D', '', cpf_proprietario)
                if len(cpf_digits) != 11:
                    messages.error(request, 'CPF deve ter 11 dígitos válidos.')
                    return render(request, 'linhas/cadastro_empresa.html')
                
                # Validar email
                from django.core.validators import validate_email
                from django.core.exceptions import ValidationError
                try:
                    validate_email(email)
                except ValidationError:
                    messages.error(request, 'Email inválido.')
                    return render(request, 'linhas/cadastro_empresa.html')
                
                # Criar empresa/cliente
                cliente = Cliente.objects.create(
                    cnpj=cnpj,
                    empresa=razao_social,
                    razao_social=razao_social,
                    fantasia=nome_fantasia or razao_social,
                    nome_dono=nome_proprietario,
                    cpf_dono=cpf_proprietario,
                    data_nascimento_dono=data_nascimento,
                    endereco_completo=endereco,
                    telefone=telefone,
                    email=email,
                    tipo_pessoa='PJ'
                )
                
                # Processar arquivo de linhas se fornecido
                linhas_importadas = 0
                if arquivo_linhas:
                    try:
                        linhas_importadas = processar_arquivo_linhas(arquivo_linhas, cliente, request.user)
                        if linhas_importadas > 0:
                            messages.success(request, f'Empresa "{razao_social}" cadastrada com sucesso! {linhas_importadas} linha{"s" if linhas_importadas > 1 else ""} importada{"s" if linhas_importadas > 1 else ""} do arquivo.')
                        else:
                            messages.success(request, f'Empresa "{razao_social}" cadastrada com sucesso!')
                            messages.warning(request, 'Nenhuma linha válida foi encontrada no arquivo.')
                    except Exception as e:
                        messages.success(request, f'Empresa "{razao_social}" cadastrada com sucesso!')
                        messages.error(request, f'Erro ao processar arquivo de linhas: {str(e)}')
                else:
                    messages.success(request, f'Empresa "{razao_social}" cadastrada com sucesso!')
                
                return redirect('linhas:cadastro_empresa')
                
        except Exception as e:
            messages.error(request, f'Erro ao cadastrar empresa: {str(e)}')
    
    return render(request, 'linhas/cadastro_empresa.html')


@login_required
def listar_empresas(request):
    """
    View para listar empresas cadastradas
    """
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Buscar todas as empresas
    empresas_qs = Cliente.objects.filter(tipo_pessoa='PJ').order_by('razao_social')
    
    # Filtros opcionais
    search = request.GET.get('search', '').strip()
    if search:
        empresas_qs = empresas_qs.filter(
            Q(razao_social__icontains=search) |
            Q(fantasia__icontains=search) |
            Q(cnpj__icontains=search) |
            Q(nome_dono__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Paginação
    paginator = Paginator(empresas_qs, 10)  # 10 empresas por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'total_empresas': empresas_qs.count(),
    }
    
    return render(request, 'linhas/listar_empresas.html', context)


def processar_arquivo_linhas(arquivo, cliente, usuario):
    """
    Processa arquivo Excel/CSV com linhas e cria registros no sistema
    """
    import csv
    import io
    from decimal import Decimal, InvalidOperation
    
    linhas_criadas = 0
    
    try:
        # Determinar tipo de arquivo e processar
        if arquivo.name.endswith('.csv'):
            # Processar CSV
            arquivo_texto = io.StringIO(arquivo.read().decode('utf-8'))
            csv_reader = csv.DictReader(arquivo_texto)
            
            for row in csv_reader:
                linha_criada = criar_linha_do_arquivo(row, cliente, usuario)
                if linha_criada:
                    linhas_criadas += 1
                    
        elif arquivo.name.endswith('.xlsx') or arquivo.name.endswith('.xls'):
            # Processar Excel
            try:
                import openpyxl
                from openpyxl import load_workbook
                
                # Carregar workbook do arquivo
                workbook = load_workbook(arquivo)
                worksheet = workbook.active
                
                # Assumir que a primeira linha contém os cabeçalhos
                headers = []
                for cell in worksheet[1]:
                    headers.append(cell.value.lower() if cell.value else '')
                
                # Processar linhas de dados
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(row):  # Pular linhas vazias
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers) and headers[i]:
                                row_dict[headers[i]] = value
                        
                        linha_criada = criar_linha_do_arquivo(row_dict, cliente, usuario)
                        if linha_criada:
                            linhas_criadas += 1
                            
            except ImportError:
                raise Exception("Biblioteca openpyxl não está instalada. Instale com: pip install openpyxl")
                
    except Exception as e:
        raise Exception(f"Erro ao processar arquivo: {str(e)}")
    
    return linhas_criadas


def criar_linha_do_arquivo(row_dict, cliente, usuario):
    """
    Cria uma linha no sistema baseado nos dados do arquivo
    """
    try:
        # Mapear campos do arquivo para campos do modelo
        numero = str(row_dict.get('numero') or row_dict.get('number') or row_dict.get('telefone') or '').strip()
        iccid = str(row_dict.get('iccid') or row_dict.get('chip') or '').strip()
        tipo_plano = str(row_dict.get('tipo_plano') or row_dict.get('plano') or row_dict.get('tipo') or '').strip()
        valor_plano_str = str(row_dict.get('valor_plano') or row_dict.get('valor') or row_dict.get('price') or '0').strip()
        rp = str(row_dict.get('rp') or row_dict.get('codigo_rp') or '').strip()
        acao = str(row_dict.get('acao') or row_dict.get('action') or 'ESTOQUE').strip().upper()
        
        # Validações básicas
        if not numero or len(numero) < 10:
            return False
            
        # Verificar se a linha já existe
        if Linha.objects.filter(numero=numero).exists():
            return False
        
        # Converter valor do plano
        try:
            valor_plano_str = valor_plano_str.replace(',', '.').replace('R$', '').strip()
            valor_plano = Decimal(valor_plano_str) if valor_plano_str else Decimal('0')
        except (InvalidOperation, ValueError):
            valor_plano = Decimal('0')
        
        # Mapear tipo de plano se necessário
        if tipo_plano and tipo_plano not in [choice[0] for choice in Linha.TIPO_CHOICES]:
            # Tentar encontrar correspondência parcial
            for choice_key, choice_label in Linha.TIPO_CHOICES:
                if any(word in tipo_plano.upper() for word in choice_key.split('_')):
                    tipo_plano = choice_key
                    break
            else:
                # Se não encontrou, usar o primeiro disponível
                tipo_plano = Linha.TIPO_CHOICES[0][0]
        
        # Validar ação
        if acao not in [choice[0] for choice in Linha.ACAO_CHOICES]:
            acao = 'ESTOQUE'
        
        # Criar linha
        linha = Linha.objects.create(
            numero=numero,
            iccid=iccid,
            cnpj=cliente.cnpj,
            empresa=cliente.empresa,
            rp=rp,
            tipo_plano=tipo_plano or Linha.TIPO_CHOICES[0][0],
            valor_plano=valor_plano,
            taxa_manutencao=Decimal('0'),
            acao=acao,
            ativa=True,
            observacoes=f'Importado automaticamente do arquivo via cadastro de empresa',
            criado_por=usuario,
            cliente=cliente
        )
        
        return True
        
    except Exception as e:
        # Log do erro para debug, mas não interrompe o processo
        print(f"Erro ao criar linha: {e}")
        return False


@login_required
def preview_arquivo_linhas(request):
    """
    AJAX view para fazer preview do arquivo de linhas antes do upload
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'})
    
    arquivo = request.FILES.get('arquivo')
    if not arquivo:
        return JsonResponse({'success': False, 'error': 'Nenhum arquivo enviado'})
    
    try:
        linhas_preview = []
        
        if arquivo.name.endswith('.csv'):
            # Processar CSV
            import csv
            import io
            
            arquivo_texto = io.StringIO(arquivo.read().decode('utf-8'))
            csv_reader = csv.DictReader(arquivo_texto)
            
            for i, row in enumerate(csv_reader):
                if i >= 10:  # Limitar preview a 10 linhas
                    break
                    
                linhas_preview.append({
                    'numero': row.get('numero') or row.get('number') or '',
                    'iccid': row.get('iccid') or row.get('chip') or '',
                    'tipo_plano': row.get('tipo_plano') or row.get('plano') or '',
                    'valor_plano': row.get('valor_plano') or row.get('valor') or '',
                    'rp': row.get('rp') or '',
                    'acao': row.get('acao') or row.get('action') or 'ESTOQUE'
                })
                
        elif arquivo.name.endswith('.xlsx') or arquivo.name.endswith('.xls'):
            # Processar Excel
            try:
                import openpyxl
                from openpyxl import load_workbook
                
                workbook = load_workbook(arquivo)
                worksheet = workbook.active
                
                # Obter cabeçalhos
                headers = []
                for cell in worksheet[1]:
                    headers.append(str(cell.value).lower() if cell.value else '')
                
                # Processar linhas (máximo 10 para preview)
                for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
                    if i >= 10 or not any(row):
                        break
                        
                    row_dict = {}
                    for j, value in enumerate(row):
                        if j < len(headers) and headers[j]:
                            row_dict[headers[j]] = str(value) if value is not None else ''
                    
                    linhas_preview.append({
                        'numero': row_dict.get('numero') or row_dict.get('number') or '',
                        'iccid': row_dict.get('iccid') or row_dict.get('chip') or '',
                        'tipo_plano': row_dict.get('tipo_plano') or row_dict.get('plano') or '',
                        'valor_plano': row_dict.get('valor_plano') or row_dict.get('valor') or '',
                        'rp': row_dict.get('rp') or '',
                        'acao': row_dict.get('acao') or row_dict.get('action') or 'ESTOQUE'
                    })
                    
            except ImportError:
                return JsonResponse({'success': False, 'error': 'Biblioteca openpyxl não está instalada. Use arquivos CSV.'})
        else:
            return JsonResponse({'success': False, 'error': 'Formato de arquivo não suportado. Use .csv, .xlsx ou .xls'})
        
        return JsonResponse({
            'success': True,
            'linhas': linhas_preview,
            'total_preview': len(linhas_preview),
            'arquivo_nome': arquivo.name
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erro ao processar arquivo: {str(e)}'})


@login_required  
def processar_lista_estoque(request):
    """
    View específica para processar o arquivo 'Lista Estoque.xlsx'
    """
    if request.method == 'POST':
        arquivo = request.FILES.get('lista_estoque')
        empresa_id = request.POST.get('empresa_id')
        
        if not arquivo:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return redirect('linhas:listar_empresas')
            
        if not empresa_id:
            messages.error(request, 'Empresa não especificada.')
            return redirect('linhas:listar_empresas')
            
        try:
            cliente = Cliente.objects.get(id=empresa_id)
            linhas_importadas = processar_arquivo_linhas(arquivo, cliente, request.user)
            
            if linhas_importadas > 0:
                messages.success(request, f'{linhas_importadas} linha{"s" if linhas_importadas > 1 else ""} importada{"s" if linhas_importadas > 1 else ""} com sucesso para {cliente.empresa}!')
            else:
                messages.warning(request, 'Nenhuma linha válida foi encontrada no arquivo.')
                
        except Cliente.DoesNotExist:
            messages.error(request, 'Empresa não encontrada.')
        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            
    # Verificar se veio das configurações
    if request.META.get('HTTP_REFERER', '').endswith('configuracoes/'):
        return redirect('linhas:configuracoes')
    return redirect('linhas:listar_empresas')

